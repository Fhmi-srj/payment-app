"""
Import santri data from Excel spreadsheet into MySQL payment_app database.
Maps spreadsheet columns to the santris table schema.
"""

import openpyxl
import mysql.connector
from datetime import datetime

# --- Configuration ---
EXCEL_PATH = r"c:\laragon\www\payment-app\DATA PENDAFTARAN SANTRI BARU.xlsx"
SHEET_NAME = "DATA PENDAFTARAN SANTRI BARU"

DB_CONFIG = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "",
    "database": "payment_app",
}

# Map spreadsheet "MASUK TINGKATAN" to database lembaga enum
LEMBAGA_MAP = {
    "MTs": "MA ALHIKAM",
    "MA": "MA ALHIKAM",
}

def main():
    # Load Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    
    records = []
    skipped = 0
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        vals = [cell.value for cell in row]
        
        nama = vals[1]  # Col B: Nama santri
        if not nama or not str(nama).strip():
            skipped += 1
            continue
        
        nama = str(nama).strip()
        alamat = str(vals[17]).strip() if vals[17] else None  # Col R: Alamat
        masuk_tingkatan = str(vals[38]).strip() if vals[38] else "MTs"  # Col AM: Masuk Tingkatan
        
        lembaga = LEMBAGA_MAP.get(masuk_tingkatan, "MA ALHIKAM")
        
        # Derive kelas from masuk_tingkatan (new students start at class 1)
        if masuk_tingkatan == "MTs":
            kelas = "7"  # MTs kelas 7 (equivalent to SMP kelas 1)
        elif masuk_tingkatan == "MA":
            kelas = "10"  # MA kelas 10 (equivalent to SMA kelas 1)
        else:
            kelas = ""
        
        records.append({
            "lembaga": lembaga,
            "nama": nama,
            "kelas": kelas,
            "alamat": alamat,
            "daftar_ulang": 0,
            "syahriyah": 0,
            "haflah": 0,
            "seragam": 0,
            "study_tour": 0,
            "sekolah": 0,
            "kartu_santri": 0,
            "status": "AKTIF",
        })
    
    print(f"Parsed {len(records)} santri records ({skipped} empty rows skipped)")
    
    # Connect to MySQL
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    
    # Check existing count
    cursor.execute("SELECT COUNT(*) FROM santris")
    existing_count = cursor.fetchone()[0]
    print(f"Existing santri in database: {existing_count}")
    
    # Insert records
    insert_sql = """
    INSERT INTO santris (lembaga, nama, kelas, alamat, daftar_ulang, syahriyah, haflah, seragam, study_tour, sekolah, kartu_santri, status, created_at, updated_at)
    VALUES (%(lembaga)s, %(nama)s, %(kelas)s, %(alamat)s, %(daftar_ulang)s, %(syahriyah)s, %(haflah)s, %(seragam)s, %(study_tour)s, %(sekolah)s, %(kartu_santri)s, %(status)s, %(created_at)s, %(updated_at)s)
    """
    
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    inserted = 0
    duplicates = 0
    
    for rec in records:
        # Check for duplicate by nama + lembaga
        cursor.execute("SELECT id FROM santris WHERE nama = %s AND lembaga = %s", (rec["nama"], rec["lembaga"]))
        if cursor.fetchone():
            duplicates += 1
            print(f"  SKIP (duplicate): {rec['nama']} ({rec['lembaga']})")
            continue
        
        rec["created_at"] = now
        rec["updated_at"] = now
        cursor.execute(insert_sql, rec)
        inserted += 1
    
    conn.commit()
    
    # Final count
    cursor.execute("SELECT COUNT(*) FROM santris")
    final_count = cursor.fetchone()[0]
    
    print(f"\n--- Import Summary ---")
    print(f"Total parsed:    {len(records)}")
    print(f"Inserted:        {inserted}")
    print(f"Duplicates:      {duplicates}")
    print(f"DB before:       {existing_count}")
    print(f"DB after:        {final_count}")
    
    cursor.close()
    conn.close()
    print("\nDone!")

if __name__ == "__main__":
    main()
